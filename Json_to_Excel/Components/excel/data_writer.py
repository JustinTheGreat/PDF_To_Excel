from datetime import datetime
from Components.utils.text_filters import TextFilter
from Components.utils.file_utils import FileUtils
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers

class ExcelDataWriter:
    """
    Enhanced class for writing data to Excel worksheets with support for complex data structures
    including nested lists, key-value pair lists, date formatting, and proper number formatting.
    """
    
    def __init__(self):
        """Initialize the data writer with date formatting settings."""
        self.date_format = 'yyyy-mm-dd'  # Default Excel date format

    def _try_parse_date(self, value):
        """
        Try to parse a value as a date.
        
        Args:
            value: The value to parse
            
        Returns:
            datetime object if successful, None if parsing fails
        """
        if not isinstance(value, str):
            return None
            
        # List of common date formats to try
        date_formats = [
            '%Y-%m-%d',     # 2024-04-22
            '%d-%m-%Y',     # 22-04-2024
            '%Y/%m/%d',     # 2024/04/22
            '%d/%m/%Y',     # 22/04/2024
            '%d.%m.%Y',     # 22.04.2024
            '%Y.%m.%d',     # 2024.04.22
            '%m/%d/%Y',     # 04/22/2024
            '%b %d %Y',     # Apr 22 2024
            '%B %d %Y',     # April 22 2024
            '%d %b %Y',     # 22 Apr 2024
            '%d %B %Y',     # 22 April 2024
            '%Y%m%d',       # 20240422
        ]
        
        # Try each format
        for fmt in date_formats:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
                
        return None

    def _is_numeric(self, value):
        """
        Check if a string value can be converted to a number.
        Handles integers, floats, and negative numbers.
        
        Args:
            value: The value to check
            
        Returns:
            Boolean indicating if the value is numeric
        """
        if value is None or not isinstance(value, str):
            return False
            
        # Strip whitespace
        value = value.strip()
        
        # Check if empty
        if not value:
            return False
            
        # Try to convert to float (handles integers too)
        try:
            float(value)
            return True
        except ValueError:
            return False
            
    def _convert_to_number(self, value):
        """
        Convert a string value to a numeric type (int or float).
        
        Args:
            value: The string value to convert
            
        Returns:
            int or float if conversion is successful, original value otherwise
        """
        if not isinstance(value, str):
            return value
            
        # Strip whitespace
        value = value.strip()
        
        # Try to convert
        try:
            # Try integer first
            int_val = int(value)
            # If int and float values are the same, return int
            if float(value) == int_val:
                return int_val
            # Otherwise, return float
            return float(value)
        except ValueError:
            try:
                # Try float
                return float(value)
            except ValueError:
                # Return original if both fail
                return value

    def add_data_row(self, worksheet, row_num, file_name, fields, structure_info, max_list_lengths, 
                     filter_text="", apply_value_filters=True, replace_commas=False):
        """
        Add a row of data to the worksheet with support for nested lists, key-value lists, date formatting,
        and proper number formatting.
        
        Args:
            worksheet: The worksheet to write to
            row_num: Row number to write at
            file_name: Name of the source file
            fields: Dictionary of field key-value pairs to write
            structure_info: Dictionary with structure information about the data
            max_list_lengths: Dictionary of maximum list lengths (deprecated)
            filter_text: Text to remove from filenames
            apply_value_filters: Whether to apply text filters to values
            replace_commas: Whether to replace commas with periods
        """
        # Process filename to remove extension and filter text
        display_filename = FileUtils.process_filename(file_name, filter_text)
        
        # Write the processed filename
        worksheet.cell(row=row_num, column=1, value=display_filename)
        
        # Start with column 2 (after file name column)
        current_column = 2
        
        # Process each field
        for key in sorted(structure_info['keys']):
            value = fields.get(key, "")
            nesting_structure = structure_info['nesting_structure'].get(key, [])
            
            # Check if this is a date field
            is_date_field = 'date' in key.lower()
            
            # Check if this is a key-value list field
            if 'kv_lists' in structure_info and key in structure_info['kv_lists']:
                # Handle key-value list type fields
                column_increment = self._add_key_value_list_data_with_nesting(
                    worksheet,
                    row_num,
                    current_column,
                    value,
                    structure_info['kv_lists'][key],
                    apply_value_filters,
                    replace_commas,
                    is_date_field
                )
                current_column += column_increment
            
            # Handle the different value types for regular lists
            elif nesting_structure:
                # This field might have nested lists
                column_increment = self._add_nested_data(
                    worksheet, 
                    row_num, 
                    current_column, 
                    value, 
                    nesting_structure, 
                    apply_value_filters,
                    replace_commas,
                    is_date_field
                )
                current_column += column_increment
            else:
                # This field has a single value or is not a list
                cell = worksheet.cell(row=row_num, column=current_column)
                
                # Set the value (single value or first item of a list)
                if isinstance(value, list) and value:
                    value_to_set = value[0]
                else:
                    value_to_set = value
                
                # Apply text filtering if needed
                if isinstance(value_to_set, str):
                    # Apply comma replacement first
                    if replace_commas:
                        value_to_set = TextFilter.replace_commas_with_periods(value_to_set)
                    
                    # Apply unit removal if needed
                    if apply_value_filters:
                        value_to_set = TextFilter.remove_units(value_to_set)
                    
                    # Handle date formatting
                    if is_date_field:
                        date_value = self._try_parse_date(value_to_set)
                        if date_value:
                            cell.value = date_value
                            cell.number_format = self.date_format
                            current_column += 1
                            continue
                    
                    # Handle numeric values - convert to actual numbers
                    if self._is_numeric(value_to_set):
                        value_to_set = self._convert_to_number(value_to_set)
                        # Apply general number format for numbers
                        cell.number_format = numbers.FORMAT_GENERAL
                
                cell.value = value_to_set
                current_column += 1
    
    def _add_key_value_list_data_with_nesting(self, worksheet, row_num, start_column, 
                                            value, kv_list_info, apply_value_filters, 
                                            replace_commas, is_date_field):
        """
        Add key-value list data to a worksheet row with support for nested objects, date formatting,
        and proper number formatting.
        
        Args:
            worksheet: The worksheet to add data to
            row_num: The row number
            start_column: The starting column
            value: The value (list of dictionaries)
            kv_list_info: Information about the key-value list structure
            apply_value_filters: Whether to apply text filters to values
            replace_commas: Whether to replace commas with periods
            is_date_field: Whether this field should be treated as a date
        
        Returns:
            The number of columns used
        """
        # Get the list of unique top-level keys
        ordered_keys = kv_list_info['unique_keys']
        nested_structure = kv_list_info.get('nested_structure', {})
        
        # Calculate total columns needed
        total_columns = 0
        for key in ordered_keys:
            if key in nested_structure:
                # Count columns for nested paths
                total_columns += len(nested_structure[key]['paths'])
            else:
                # One column for regular key
                total_columns += 1
        
        # Initialize all cells to empty
        for col in range(start_column, start_column + total_columns):
            worksheet.cell(row=row_num, column=col, value="")
        
        # Handle if value is not a list or is empty
        if not isinstance(value, list) or not value:
            return total_columns
        
        # Get the first item in the list
        first_item = value[0]
        if not isinstance(first_item, dict):
            return total_columns
        
        current_column = start_column
        
        # Process each top-level key
        for key in ordered_keys:
            if key in first_item:
                item_value = first_item[key]
                
                if key in nested_structure and isinstance(item_value, dict):
                    # Handle nested object
                    flattened = {}
                    self._flatten_object(item_value, "", flattened)
                    
                    # Write each property value to its column
                    for path, path_info in nested_structure[key]['paths'].items():
                        # Get the property name (last part of path)
                        prop_name = path.split('.')[-1]
                        
                        # Get value or empty string if not found
                        prop_value = flattened.get(prop_name, "")
                        
                        cell = worksheet.cell(row=row_num, column=current_column)
                        
                        # Apply filters if needed
                        if isinstance(prop_value, str):
                            # Apply comma replacement first
                            if replace_commas:
                                prop_value = TextFilter.replace_commas_with_periods(prop_value)
                            
                            # Apply unit removal if needed
                            if apply_value_filters:
                                prop_value = TextFilter.remove_units(prop_value)
                            
                            # Handle date formatting for nested properties
                            if is_date_field:
                                date_value = self._try_parse_date(prop_value)
                                if date_value:
                                    cell.value = date_value
                                    cell.number_format = self.date_format
                                    current_column += 1
                                    continue
                            
                            # Handle numeric values
                            if self._is_numeric(prop_value):
                                prop_value = self._convert_to_number(prop_value)
                                # Apply general number format
                                cell.number_format = numbers.FORMAT_GENERAL
                        
                        cell.value = prop_value
                        current_column += 1
                else:
                    # Handle regular key
                    cell = worksheet.cell(row=row_num, column=current_column)
                    
                    # Apply filters if needed
                    if isinstance(item_value, str):
                        # Apply comma replacement first
                        if replace_commas:
                            item_value = TextFilter.replace_commas_with_periods(item_value)
                        
                        # Apply unit removal if needed
                        if apply_value_filters:
                            item_value = TextFilter.remove_units(item_value)
                        
                        # Handle date formatting
                        if is_date_field:
                            date_value = self._try_parse_date(item_value)
                            if date_value:
                                cell.value = date_value
                                cell.number_format = self.date_format
                                current_column += 1
                                continue
                        
                        # Handle numeric values
                        if self._is_numeric(item_value):
                            item_value = self._convert_to_number(item_value)
                            # Apply general number format
                            cell.number_format = numbers.FORMAT_GENERAL
                    
                    cell.value = item_value
                    current_column += 1
            else:
                # Key not in item, skip columns
                if key in nested_structure:
                    # Skip columns for nested properties
                    current_column += len(nested_structure[key]['paths'])
                else:
                    # Skip one column for regular key
                    current_column += 1
        
        return total_columns

    def _add_nested_data(self, worksheet, row_num, start_column, value, dimensions, 
                        apply_value_filters, replace_commas, is_date_field):
        """
        Add nested data to a worksheet row with date support and proper number formatting.
        
        Args:
            worksheet: The worksheet to add data to
            row_num: The row number
            start_column: The starting column
            value: The value (possibly nested list)
            dimensions: List of dimensions for the nested structure
            apply_value_filters: Whether to apply text filters to values
            replace_commas: Whether to replace commas with periods
            is_date_field: Whether this field should be treated as a date
        
        Returns:
            The number of columns used
        """
        if not dimensions:
            cell = worksheet.cell(row=row_num, column=start_column)
            
            if isinstance(value, str):
                # Apply comma replacement first
                if replace_commas:
                    value = TextFilter.replace_commas_with_periods(value)
                
                # Apply unit removal if needed
                if apply_value_filters:
                    value = TextFilter.remove_units(value)
                
                # Handle date formatting
                if is_date_field:
                    date_value = self._try_parse_date(value)
                    if date_value:
                        cell.value = date_value
                        cell.number_format = self.date_format
                        return 1
                
                # Handle numeric values
                if self._is_numeric(value):
                    value = self._convert_to_number(value)
                    # Apply general number format
                    cell.number_format = numbers.FORMAT_GENERAL
            
            cell.value = value
            return 1
        
        # Calculate total columns needed
        total_columns = self._calculate_total_columns(dimensions)
        
        # Initialize all cells to empty
        for col in range(start_column, start_column + total_columns):
            worksheet.cell(row=row_num, column=col, value="")
        
        # Flatten the nested list structure
        flattened_values = []
        self._flatten_nested_list(value, flattened_values, dimensions)
        
        # Add values to cells
        for i, item in enumerate(flattened_values):
            if i < total_columns:
                cell = worksheet.cell(row=row_num, column=start_column + i)
                
                if isinstance(item, str):
                    # Apply comma replacement first
                    if replace_commas:
                        item = TextFilter.replace_commas_with_periods(item)
                    
                    # Apply unit removal if needed
                    if apply_value_filters:
                        item = TextFilter.remove_units(item)
                    
                    # Handle date formatting for nested items
                    if is_date_field:
                        date_value = self._try_parse_date(item)
                        if date_value:
                            cell.value = date_value
                            cell.number_format = self.date_format
                            continue
                    
                    # Handle numeric values
                    if self._is_numeric(item):
                        item = self._convert_to_number(item)
                        # Apply general number format
                        cell.number_format = numbers.FORMAT_GENERAL
                
                cell.value = item
        
        return total_columns

    def _calculate_total_columns(self, dimensions):
        """
        Calculate the total number of columns needed for a nested structure.
        
        Args:
            dimensions: List of dimensions at each nesting level [d1, d2, d3, ...]
        
        Returns:
            Total number of columns needed
        """
        if not dimensions:
            return 1
        
        # Multiply all dimensions together
        total = 1
        for dim in dimensions:
            total *= max(1, dim)  # Ensure at least 1 column even for empty dimensions
        
        return total
    
    def _flatten_nested_list(self, value, result, dimensions, current_dim=0):
        """
        Recursively flatten a nested list structure.
        
        Args:
            value: The value to flatten (may be a nested list)
            result: List to store flattened values
            dimensions: List of dimensions for the nested structure
            current_dim: Current dimension being processed
        """
        if not isinstance(value, list):
            # Base case: not a list, add the value
            result.append(value)
            return
        
        # If we've reached the end of our dimensions but still have a list
        if current_dim >= len(dimensions):
            # Just add the first item if available
            if value:
                result.append(value[0])
            else:
                result.append("")
            return
        
        # Current dimension size
        dim_size = dimensions[current_dim]
        
        # Process each item in the current dimension
        for i in range(dim_size):
            if i < len(value):
                # Recurse with the nested item
                self._flatten_nested_list(value[i], result, dimensions, current_dim + 1)
            else:
                # Fill in with blanks for missing items
                # Calculate how many empty slots to add
                if current_dim < len(dimensions) - 1:
                    empties_to_add = self._calculate_total_columns(dimensions[current_dim + 1:])
                else:
                    empties_to_add = 1
                
                for _ in range(empties_to_add):
                    result.append("")

    def _flatten_object(self, obj, prefix="", result=None):
        """
        Flatten a nested object into a dictionary with path keys.
        
        Args:
            obj: The object to flatten
            prefix: Current path prefix
            result: Dictionary to store results (created if None)
            
        Returns:
            Flattened dictionary with path keys
        """
        if result is None:
            result = {}
        
        if not isinstance(obj, dict):
            result[prefix] = obj
            return result
        
        for key, value in obj.items():
            path = f"{prefix}.{key}" if prefix else key
            
            if isinstance(value, dict):
                # Recursively flatten nested dictionaries
                self._flatten_object(value, path, result)
            else:
                # Add leaf value
                result[path] = value
        
        return result