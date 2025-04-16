from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExcelFormatter:
    """
    Enhanced class for handling Excel formatting operations with support for key-value lists.
    """
    
    def __init__(self):
        """Initialize Excel formatter with default styling."""
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_style = {
            'font': Font(bold=True),
            'fill': PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"),
            'border': self.thin_border
        }
        
        self.subtitle_style = {
            'font': Font(bold=True, italic=True),
            'fill': PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"),
            'border': self.thin_border
        }
    
    def apply_cell_style(self, cell, style_dict):
        """Apply a dictionary of styles to a cell."""
        for attr, value in style_dict.items():
            setattr(cell, attr, value)
    
    def sanitize_sheet_name(self, sheet_name, max_length=31):
        """
        Sanitize a sheet name for use in Excel.
        
        Args:
            sheet_name: The original sheet name to sanitize
            max_length: Maximum length for Excel sheet names (default: 31)
            
        Returns:
            Sanitized sheet name
        """
        # Remove invalid characters
        safe_name = ''.join(c for c in sheet_name if c not in '\\/:*?[]')
        
        # Truncate to maximum length
        safe_name = safe_name[:max_length]
        
        return safe_name
    
    def setup_headers(self, worksheet, structure_info):
        """Set up the headers for a worksheet with support for nested lists and key-value lists."""
        # Set up the filename header
        filename_header = worksheet.cell(row=1, column=1, value="File Name")
        self.apply_cell_style(filename_header, self.header_style)
        
        # Determine the number of subtitle rows needed
        max_nesting_level = 0
        for key in structure_info['keys']:
            if key in structure_info['nesting_depth']:
                max_nesting_level = max(max_nesting_level, structure_info['nesting_depth'][key])
        
        num_subtitle_rows = max_nesting_level if max_nesting_level > 0 else 0
        
        # Add subtitle rows if needed
        for row in range(2, 2 + num_subtitle_rows):
            subtitle_cell = worksheet.cell(row=row, column=1, value="")
            self.apply_cell_style(subtitle_cell, self.subtitle_style)
        
        # Set up field headers
        current_column = 2
        for key in sorted(structure_info['keys']):
            # Check if this is a key-value list field
            if 'kv_lists' in structure_info and key in structure_info['kv_lists']:
                # Handle key-value list type fields
                current_column = self._setup_key_value_list_headers(
                    worksheet,
                    current_column,
                    key,
                    structure_info['kv_lists'][key],
                    num_subtitle_rows
                )
                continue
            
            # Handle regular nested lists
            nesting_depth = structure_info['nesting_depth'].get(key, 0)
            nesting_structure = structure_info['nesting_structure'].get(key, [])
            
            # Calculate total columns needed for this field
            total_columns = self._calculate_total_columns(nesting_structure)
            
            # Set the header (key)
            header_cell = worksheet.cell(row=1, column=current_column, value=key)
            self.apply_cell_style(header_cell, self.header_style)
            
            if total_columns > 1:
                # This field has multiple items - needs subtitles
                # First, merge the header cell across all the items
                merge_end_column = current_column + total_columns - 1
                worksheet.merge_cells(
                    start_row=1, 
                    start_column=current_column, 
                    end_row=1, 
                    end_column=merge_end_column
                )
                
                # Center the merged header
                header_cell.alignment = Alignment(horizontal='center')
                
                # Generate hierarchical subtitles
                self._create_hierarchical_subtitles(
                    worksheet, 
                    key, 
                    current_column, 
                    nesting_structure, 
                    2,  # Start at row 2
                    num_subtitle_rows
                )
                
                current_column += total_columns
            else:
                # This field has a single value or is not a list
                if structure_info['needs_subtitles']:
                    # If other fields have subtitles, add blank subtitle cells for consistency
                    for row in range(2, 2 + num_subtitle_rows):
                        subtitle_cell = worksheet.cell(row=row, column=current_column, value="")
                        self.apply_cell_style(subtitle_cell, self.subtitle_style)
                
                current_column += 1
    
    def _setup_key_value_list_headers(self, worksheet, start_column, parent_key, kv_list_info, num_subtitle_rows):
        """
        Set up headers for a key-value list field.
        
        Args:
            worksheet: The worksheet to add headers to
            start_column: Starting column for the headers
            parent_key: The parent field key
            kv_list_info: Information about the key-value list structure
            num_subtitle_rows: Number of subtitle rows available
            
        Returns:
            Next column position after setting up headers
        """
        # Get the list of unique keys in the dictionary items
        unique_keys = kv_list_info['unique_keys']
        total_columns = len(unique_keys)
        
        # Create the parent header merged across all columns
        header_cell = worksheet.cell(row=1, column=start_column, value=parent_key)
        self.apply_cell_style(header_cell, self.header_style)
        
        if total_columns > 1:
            # Merge the parent header across all columns
            merge_end_column = start_column + total_columns - 1
            worksheet.merge_cells(
                start_row=1,
                start_column=start_column,
                end_row=1,
                end_column=merge_end_column
            )
            
            # Center the merged header
            header_cell.alignment = Alignment(horizontal='center')
            
            # Create subtitles for each key in the key-value list
            for i, key in enumerate(unique_keys):
                col = start_column + i
                subtitle_cell = worksheet.cell(row=2, column=col, value=key)
                self.apply_cell_style(subtitle_cell, self.subtitle_style)
                
                # Add empty subtitle cells for any remaining subtitle rows
                for row in range(3, 2 + num_subtitle_rows):
                    empty_cell = worksheet.cell(row=row, column=col, value="")
                    self.apply_cell_style(empty_cell, self.subtitle_style)
        else:
            # Only one key, use it as subtitle if there are subtitle rows
            if num_subtitle_rows > 0:
                subtitle_cell = worksheet.cell(row=2, column=start_column, value=unique_keys[0])
                self.apply_cell_style(subtitle_cell, self.subtitle_style)
                
                # Add empty subtitle cells for any remaining subtitle rows
                for row in range(3, 2 + num_subtitle_rows):
                    empty_cell = worksheet.cell(row=row, column=start_column, value="")
                    self.apply_cell_style(empty_cell, self.subtitle_style)
        
        return start_column + total_columns
    
    def _calculate_total_columns(self, dimensions):
        """
        Calculate the total number of columns needed for a nested structure.
        
        Args:
            dimensions: List of dimensions at each nesting level [d1, d2, d3, ...]
        
        Returns:
            Total number of columns needed
        """
        if not dimensions:
            return 1
        
        # Multiply all dimensions together
        total = 1
        for dim in dimensions:
            total *= max(1, dim)  # Ensure at least 1 column even for empty dimensions
        
        return total
    
    def _create_hierarchical_subtitles(self, worksheet, key, start_column, dimensions, start_row, max_rows):
        """
        Create hierarchical subtitles for nested lists.
        
        Args:
            worksheet: The worksheet to add subtitles to
            key: The field key (for naming)
            start_column: Starting column for the subtitles
            dimensions: List of dimensions at each nesting level [d1, d2, d3, ...]
            start_row: Starting row for the subtitles
            max_rows: Maximum number of subtitle rows
        """
        if not dimensions:
            return
        
        # Create a recursive function to generate subtitles
        def create_subtitles(level, prefix, col_start, col_span, row):
            if level >= len(dimensions) or row > max_rows + 1:
                return
            
            dim = dimensions[level]
            if dim <= 0:
                # Handle empty level
                dim = 1
            
            # Calculate column span for each item at this level
            item_span = col_span // dim
            
            for i in range(dim):
                # Calculate column range for this item
                item_start = col_start + (i * item_span)
                item_end = item_start + item_span - 1
                
                # Create the subtitle for this item
                if prefix:
                    subtitle = f"{prefix} - #{i+1}"
                else:
                    subtitle = f"{key} - #{i+1}"
                
                # Create and merge the subtitle cell
                subtitle_cell = worksheet.cell(row=row, column=item_start, value=subtitle)
                self.apply_cell_style(subtitle_cell, self.subtitle_style)
                
                if item_span > 1:
                    worksheet.merge_cells(
                        start_row=row,
                        start_column=item_start,
                        end_row=row,
                        end_column=item_end
                    )
                    subtitle_cell.alignment = Alignment(horizontal='center')
                
                # Recurse to next level with updated prefix
                if level < len(dimensions) - 1:
                    create_subtitles(level + 1, subtitle, item_start, item_span, row + 1)
        
        # Start the recursive subtitle creation
        create_subtitles(0, "", start_column, self._calculate_total_columns(dimensions), start_row)
    
    def get_column_count(self, structure_info):
        """Calculate the total number of columns needed based on structure info including key-value lists."""
        count = 1  # Start with 1 for the filename column
        
        for key in structure_info['keys']:
            # Handle key-value list fields
            if 'kv_lists' in structure_info and key in structure_info['kv_lists']:
                # For key-value lists, count the number of unique keys
                count += len(structure_info['kv_lists'][key]['unique_keys'])
            else:
                # Handle regular nested lists
                nesting_structure = structure_info['nesting_structure'].get(key, [])
                if nesting_structure:
                    count += self._calculate_total_columns(nesting_structure)
                else:
                    count += 1
                
        return count
    
    def adjust_column_widths(self, worksheet, num_columns, last_row):
        """Adjust column widths based on content."""
        for col_idx in range(1, num_columns + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            
            # Check all rows
            for row in range(1, last_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value:
                    text_length = len(str(cell.value))
                    max_length = max(max_length, text_length)
            
            # Set column width (with some padding)
            if max_length > 0:
                adjusted_width = max_length + 2  # Add padding
                worksheet.column_dimensions[column].width = adjusted_width
                # Add to ExcelFormatter class

    def _setup_key_value_list_headers_with_nesting(self, worksheet, start_column, parent_key, 
                                                kv_list_info, start_row, max_rows):
        """
        Set up headers for a key-value list field with support for nested objects.
        
        Args:
            worksheet: The worksheet to add headers to
            start_column: Starting column for the headers
            parent_key: The parent field key
            kv_list_info: Information about the key-value list structure
            start_row: Starting row for the headers
            max_rows: Maximum number of subtitle rows
            
        Returns:
            Next column position after setting up headers
        """
        # Get the list of unique top-level keys
        unique_keys = kv_list_info['unique_keys']
        nested_structure = kv_list_info.get('nested_structure', {})
        
        # Create the parent header
        header_cell = worksheet.cell(row=start_row, column=start_column, value=parent_key)
        self.apply_cell_style(header_cell, self.header_style)
        
        current_column = start_column
        
        # Create a function to recursively generate headers
        def create_nested_headers(key, prefix, col, row, nested_obj=None):
            nonlocal current_column
            
            # Check if this key has nested objects
            if key in nested_structure and nested_obj is None:
                # This is a top-level key with nested objects
                nested_obj = nested_structure[key]
                
                # Create subtitle for this key
                subtitle_cell = worksheet.cell(row=row, column=col, value=key)
                self.apply_cell_style(subtitle_cell, self.subtitle_style)
                
                # Calculate columns needed for nested properties
                nested_paths = nested_obj['paths']
                nested_columns = len(nested_paths)
                
                if nested_columns > 1:
                    # Merge the subtitle cell across nested columns
                    worksheet.merge_cells(
                        start_row=row,
                        start_column=col,
                        end_row=row,
                        end_column=col + nested_columns - 1
                    )
                    subtitle_cell.alignment = Alignment(horizontal='center')
                
                # Create subtitles for nested properties
                next_row = row + 1
                if next_row <= max_rows:
                    for path, path_info in nested_paths.items():
                        # Get the property name (last part of path)
                        prop_name = path.split('.')[-1]
                        subtitle_cell = worksheet.cell(row=next_row, column=current_column, value=prop_name)
                        self.apply_cell_style(subtitle_cell, self.subtitle_style)
                        current_column += 1
                else:
                    # No space for nested subtitles, just move column pointer
                    current_column += nested_columns
                
                return nested_columns
            else:
                # Regular key without nesting
                subtitle_cell = worksheet.cell(row=row, column=col, value=key)
                self.apply_cell_style(subtitle_cell, self.subtitle_style)
                current_column += 1
                return 1
        
        # Create subtitles for each key in the key-value list
        for key in unique_keys:
            create_nested_headers(key, parent_key, current_column, start_row + 1)
        
        # Merge the parent header across all columns
        total_columns = current_column - start_column
        if total_columns > 1:
            worksheet.merge_cells(
                start_row=start_row,
                start_column=start_column,
                end_row=start_row,
                end_column=start_column + total_columns - 1
            )
            header_cell.alignment = Alignment(horizontal='center')
        
        return current_column

    # Replace the existing _setup_key_value_list_headers with a call to the enhanced version
    def _setup_key_value_list_headers(self, worksheet, start_column, parent_key, 
                                    kv_list_info, num_subtitle_rows):
        """
        Set up headers for a key-value list field.
        """
        # Call the enhanced version with nesting support
        return self._setup_key_value_list_headers_with_nesting(
            worksheet, 
            start_column, 
            parent_key, 
            kv_list_info,
            1,  # Start at row 1 for the main header
            num_subtitle_rows + 1  # +1 because row 1 is the main header
        )
    # Add these methods to the ExcelFormatter class

    def _setup_key_value_list_headers_with_nesting(self, worksheet, start_column, parent_key, 
                                                kv_list_info, num_subtitle_rows):
        """
        Set up headers for a key-value list field with support for nested objects.
        
        Args:
            worksheet: The worksheet to add headers to
            start_column: Starting column for the headers
            parent_key: The parent field key
            kv_list_info: Information about the key-value list structure
            num_subtitle_rows: Maximum number of subtitle rows available
            
        Returns:
            Next column position after setting up headers
        """
        # Get the list of unique top-level keys
        unique_keys = kv_list_info['unique_keys']
        nested_structure = kv_list_info.get('nested_structure', {})
        
        # Create the parent header
        header_cell = worksheet.cell(row=1, column=start_column, value=parent_key)
        self.apply_cell_style(header_cell, self.header_style)
        
        current_column = start_column
        total_columns = 0
        
        # Process each top-level key
        for key in unique_keys:
            if key in nested_structure:
                # This is a key with nested objects
                nested_obj = nested_structure[key]
                nested_paths = nested_obj['paths']
                
                # Get flat list of all paths (will be shown as separate columns)
                path_keys = list(nested_paths.keys())
                
                # Create a header for this key in the first subtitle row
                subtitle_cell = worksheet.cell(row=2, column=current_column, value=key)
                self.apply_cell_style(subtitle_cell, self.subtitle_style)
                
                # If there are multiple paths, merge across all columns
                if len(path_keys) > 1:
                    worksheet.merge_cells(
                        start_row=2,
                        start_column=current_column,
                        end_row=2,
                        end_column=current_column + len(path_keys) - 1
                    )
                    subtitle_cell.alignment = Alignment(horizontal='center')
                
                # Add subtitle rows for nested properties
                for path_idx, path in enumerate(path_keys):
                    # Get the property name (last part of path)
                    prop_name = path.split('.')[-1]
                    
                    # Create subtitle cell for this property
                    col = current_column + path_idx
                    prop_cell = worksheet.cell(row=3, column=col, value=prop_name)
                    self.apply_cell_style(prop_cell, self.subtitle_style)
                    
                    # Fill remaining subtitle rows (if any) with empty styled cells
                    for row in range(4, 2 + num_subtitle_rows):
                        empty_cell = worksheet.cell(row=row, column=col, value="")
                        self.apply_cell_style(empty_cell, self.subtitle_style)
                
                # Update column counter
                current_column += len(path_keys)
                total_columns += len(path_keys)
            else:
                # Regular key without nesting
                subtitle_cell = worksheet.cell(row=2, column=current_column, value=key)
                self.apply_cell_style(subtitle_cell, self.subtitle_style)
                
                # Fill remaining subtitle rows (if any) with empty styled cells
                for row in range(3, 2 + num_subtitle_rows):
                    empty_cell = worksheet.cell(row=row, column=current_column, value="")
                    self.apply_cell_style(empty_cell, self.subtitle_style)
                
                current_column += 1
                total_columns += 1
        
        # Merge the parent header across all columns
        if total_columns > 1:
            worksheet.merge_cells(
                start_row=1,
                start_column=start_column,
                end_row=1,
                end_column=start_column + total_columns - 1
            )
            header_cell.alignment = Alignment(horizontal='center')
        
        return current_column

    # This method should be modified to call our new implementation
    def _setup_key_value_list_headers(self, worksheet, start_column, parent_key, 
                                    kv_list_info, num_subtitle_rows):
        """
        Set up headers for a key-value list field.
        
        Args:
            worksheet: The worksheet to add headers to
            start_column: Starting column for the headers
            parent_key: The parent field key
            kv_list_info: Information about the key-value list structure
            num_subtitle_rows: Number of subtitle rows available
            
        Returns:
            Next column position after setting up headers
        """
        # Check if there are nested objects in this key-value list
        if 'nested_structure' in kv_list_info and kv_list_info['nested_structure']:
            # Use the enhanced version with nesting support
            return self._setup_key_value_list_headers_with_nesting(
                worksheet, 
                start_column, 
                parent_key, 
                kv_list_info,
                num_subtitle_rows
            )
        
        # Otherwise, use the original implementation for simple key-value lists
        # [Keep the existing implementation from the original code]
        
        # Get the list of unique keys in the dictionary items
        unique_keys = kv_list_info['unique_keys']
        total_columns = len(unique_keys)
        
        # Create the parent header merged across all columns
        header_cell = worksheet.cell(row=1, column=start_column, value=parent_key)
        self.apply_cell_style(header_cell, self.header_style)
        
        if total_columns > 1:
            # Merge the parent header across all columns
            merge_end_column = start_column + total_columns - 1
            worksheet.merge_cells(
                start_row=1,
                start_column=start_column,
                end_row=1,
                end_column=merge_end_column
            )
            
            # Center the merged header
            header_cell.alignment = Alignment(horizontal='center')
            
            # Create subtitles for each key in the key-value list
            for i, key in enumerate(unique_keys):
                col = start_column + i
                subtitle_cell = worksheet.cell(row=2, column=col, value=key)
                self.apply_cell_style(subtitle_cell, self.subtitle_style)
                
                # Add empty subtitle cells for any remaining subtitle rows
                for row in range(3, 2 + num_subtitle_rows):
                    empty_cell = worksheet.cell(row=row, column=col, value="")
                    self.apply_cell_style(empty_cell, self.subtitle_style)
        else:
            # Only one key, use it as subtitle if there are subtitle rows
            if num_subtitle_rows > 0:
                subtitle_cell = worksheet.cell(row=2, column=start_column, value=unique_keys[0])
                self.apply_cell_style(subtitle_cell, self.subtitle_style)
                
                # Add empty subtitle cells for any remaining subtitle rows
                for row in range(3, 2 + num_subtitle_rows):
                    empty_cell = worksheet.cell(row=row, column=start_column, value="")
                    self.apply_cell_style(empty_cell, self.subtitle_style)
        
        return start_column + total_columns