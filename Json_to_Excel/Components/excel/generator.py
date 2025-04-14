import openpyxl
import traceback
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from Components.excel.formatter import ExcelFormatter
from Components.excel.data_writer import ExcelDataWriter
from Components.json.analyzer import JsonAnalyzer

class ExcelGenerator:
    def __init__(self):
        """Initialize Excel generator with default styling."""
        # Create formatter and writer helpers
        self.formatter = ExcelFormatter()
        self.data_writer = ExcelDataWriter()
    
    def create_excel_file(self, all_json_data, output_excel_path, filter_text="", 
                          apply_value_filters=True, callback=None):
        """
        Create an Excel file based on the JSON data with support for nested lists.
        
        Args:
            all_json_data: Dictionary of JSON data keyed by filename
            output_excel_path: Path where the Excel file will be saved
            filter_text: Text to remove from filenames
            apply_value_filters: Whether to apply text filters to values
            callback: Function to call with status updates (optional)
        """
        try:
            # Debug helper function
            def debug(message):
                if callback:
                    callback("debug", f"Excel Generator: {message}")
                print(message)
            
            debug(f"Starting Excel generation with {len(all_json_data)} files")
            debug(f"Output path: {output_excel_path}")
            
            # Create a new workbook
            workbook = openpyxl.Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            debug("Created new workbook")
            
            # Dictionary to track worksheets by title and their current row
            worksheets = {}  # {title: {'sheet': worksheet, 'next_row': row_number}}
            
            # First pass: analyze all reports to get complete structure information
            debug("First pass: analyzing all reports to determine complete structure")
            all_structure_info = {}  # {title: structure_info}
            
            for file_name, file_json_data in all_json_data.items():
                # Convert to list if not already
                reports_to_process = file_json_data if isinstance(file_json_data, list) else [file_json_data]
                
                for report in reports_to_process:
                    # Extract the title
                    title = None
                    if isinstance(report, dict):
                        title = report.get('title', None)
                    
                    if title is None:
                        title = f"Report_{file_name}"
                    
                    # Make a safe title for Excel worksheet
                    safe_title = self.formatter.sanitize_sheet_name(title)
                    
                    # Analyze this report's structure
                    this_structure = JsonAnalyzer.analyze_json_structure([report], False)
                    
                    # Merge with existing structure info for this title
                    if safe_title not in all_structure_info:
                        all_structure_info[safe_title] = this_structure
                    else:
                        # Merge keys
                        all_structure_info[safe_title]['keys'].update(this_structure['keys'])
                        
                        # Update nesting depth and structure if deeper
                        for key, depth in this_structure['nesting_depth'].items():
                            if key not in all_structure_info[safe_title]['nesting_depth'] or \
                               depth > all_structure_info[safe_title]['nesting_depth'][key]:
                                all_structure_info[safe_title]['nesting_depth'][key] = depth
                                all_structure_info[safe_title]['nesting_structure'][key] = \
                                    this_structure['nesting_structure'].get(key, [])
                        
                        # Update subtitle flag
                        all_structure_info[safe_title]['needs_subtitles'] = \
                            all_structure_info[safe_title]['needs_subtitles'] or \
                            this_structure['needs_subtitles']
            
            # Print structure info for debugging
            for title, structure in all_structure_info.items():
                debug(f"Structure for worksheet '{title}':")
                debug(f"  - {len(structure['keys'])} unique keys")
                debug(f"  - Needs subtitles: {structure['needs_subtitles']}")
                for key, depth in structure['nesting_depth'].items():
                    if depth > 0:
                        dimensions = structure['nesting_structure'].get(key, [])
                        debug(f"  - Field '{key}': depth={depth}, dimensions={dimensions}")
            
            # Process each file's JSON data
            total_files = len(all_json_data)
            total_reports_processed = 0
            
            for file_index, (file_name, file_json_data) in enumerate(all_json_data.items()):
                # Update progress if callback provided
                if callback:
                    callback("status", f"Processing {file_name}...")
                    callback("progress", (file_index + 1) / total_files * 100)
                
                debug(f"Processing file {file_index+1}/{total_files}: {file_name}")
                
                # Check the type of the JSON data
                if isinstance(file_json_data, dict):
                    debug(f"  File data is a dictionary with {len(file_json_data)} keys")
                    if file_json_data:
                        debug(f"  Keys: {list(file_json_data.keys())[:5]}...")
                    # Check if there's a title key which suggests it's a single report
                    if 'title' in file_json_data:
                        debug("  Found 'title' key, treating as a single report")
                        reports_to_process = [file_json_data]
                    else:
                        debug("  No 'title' key found, treating entire dictionary as a single report")
                        reports_to_process = [file_json_data]
                elif isinstance(file_json_data, list):
                    debug(f"  File data is a list with {len(file_json_data)} items")
                    if file_json_data and isinstance(file_json_data[0], dict):
                        sample_keys = list(file_json_data[0].keys())[:5]
                        debug(f"  First item keys: {sample_keys}")
                    reports_to_process = file_json_data
                else:
                    debug(f"  File data is a {type(file_json_data).__name__}, not a dict or list")
                    debug("  Wrapping in a list for processing")
                    reports_to_process = [file_json_data]
                
                debug(f"  Will process {len(reports_to_process)} reports from this file")
                
                # Process each report in the JSON data
                for report_index, report in enumerate(reports_to_process):
                    debug(f"  Processing report {report_index+1}/{len(reports_to_process)}")
                    
                    # Extract the title
                    title = None
                    if isinstance(report, dict):
                        title = report.get('title', None)
                    
                    if title is None:
                        title = f"Report_{file_name}_{report_index}"
                        debug(f"  No title found, using generated title: {title}")
                    else:
                        debug(f"  Report title: {title}")
                    
                    # Process this report
                    total_reports_processed += 1
                    
                    # Make a safe title for Excel worksheet
                    safe_title = self.formatter.sanitize_sheet_name(title)
                    
                    # Check if we already have a worksheet for this title
                    if safe_title in worksheets:
                        debug(f"  Adding to existing worksheet: {safe_title}")
                        worksheet = worksheets[safe_title]['sheet']
                        next_row = worksheets[safe_title]['next_row']
                    else:
                        # Create a new worksheet
                        worksheet = workbook.create_sheet(title=safe_title)
                        debug(f"  Created new worksheet: {safe_title}")
                        
                        # Use the complete structure info we gathered in the first pass
                        structure_info = all_structure_info[safe_title]
                        
                        # Set up the headers
                        self.formatter.setup_headers(worksheet, structure_info)
                        
                        # Determine start row based on nesting depth
                        max_nesting_level = 0
                        for key in structure_info['keys']:
                            if key in structure_info['nesting_depth']:
                                max_nesting_level = max(max_nesting_level, structure_info['nesting_depth'][key])
                        
                        next_row = 2 + max_nesting_level  # Start after header and subtitle rows
                        
                        # Store worksheet info
                        worksheets[safe_title] = {
                            'sheet': worksheet,
                            'next_row': next_row,
                            'structure_info': structure_info,
                            'column_count': self.formatter.get_column_count(structure_info)
                        }
                    
                    # Extract fields from the report
                    if isinstance(report, dict) and 'fields' in report:
                        fields = report.get('fields', {})
                        debug(f"  Using 'fields' section with {len(fields)} keys")
                    elif isinstance(report, dict):
                        fields = report
                        debug(f"  Using entire report as fields with {len(fields)} keys")
                    else:
                        debug(f"  Report is not a dictionary, it's a {type(report).__name__}")
                        fields = {}
                    
                    # Add this file's data to the worksheet
                    structure_info = worksheets[safe_title]['structure_info']
                    self.data_writer.add_data_row(
                        worksheet, 
                        next_row, 
                        file_name, 
                        fields, 
                        structure_info, 
                        {},  # No need for max_list_lengths anymore, using nesting_structure instead
                        filter_text, 
                        apply_value_filters
                    )
                    
                    # Update the next row
                    worksheets[safe_title]['next_row'] = next_row + 1
            
            # Auto-adjust column widths for all worksheets
            debug("Adjusting column widths for all worksheets")
            for title, ws_info in worksheets.items():
                worksheet = ws_info['sheet']
                last_row = ws_info['next_row'] - 1
                column_count = ws_info['column_count']
                debug(f"  Adjusting widths for worksheet '{title}' with {column_count} columns and {last_row} rows")
                self.formatter.adjust_column_widths(worksheet, column_count + 1, last_row)  # +1 for safety
            
            debug(f"All processing complete. Processed {total_reports_processed} reports from {total_files} files.")
            debug(f"Created {len(worksheets)} worksheets.")
            debug(f"Saving workbook to {output_excel_path}")
            
            # Save the workbook
            workbook.save(output_excel_path)
            
            if callback:
                callback("status", f"Excel file created successfully at {output_excel_path}")
            return True
        
        except Exception as e:
            error_message = f"Error: {str(e)}"
            stack_trace = traceback.format_exc()
            
            if callback:
                callback("status", error_message)
                callback("debug", f"EXCEPTION: {error_message}")
                callback("debug", f"STACK TRACE: {stack_trace}")
            
            print(error_message)
            print(stack_trace)
            return False